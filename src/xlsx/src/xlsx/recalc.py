#!/usr/bin/env python3
"""Excel Formula Recalculation Script"""

import os
import platform
import subprocess
from pathlib import Path

from openpyxl import load_workbook


def setup_libreoffice_macro() -> bool:
    """Setup LibreOffice macro for recalculation if not already configured.

    Returns:
        bool: True if macro setup succeeds, False otherwise
    """
    if platform.system() == "Darwin":
        macro_dir = os.path.expanduser("~/Library/Application Support/LibreOffice/4/user/basic/Standard")
    else:
        macro_dir = os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")

    macro_file = os.path.join(macro_dir, "Module1.xba")

    macro_content = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

    # Check if macro already exists with correct content
    if os.path.exists(macro_file):
        try:
            with open(macro_file, "r") as f:
                existing_content = f.read()
                if "RecalculateAndSave" in existing_content:
                    # Macro already exists, only overwrite if content differs
                    if existing_content.strip() == macro_content.strip():
                        return True
        except Exception:
            pass  # If reading fails, proceed to recreate

    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        with open(macro_file, "w") as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename: str, timeout: int = 30) -> dict:
    """Recalculate formulas in Excel file and report any errors.

    Args:
        filename: Path to the Excel file
        timeout: Maximum time in seconds to wait for recalculation

    Returns:
        dict: Result containing status, error details, or error message
    """
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except subprocess.TimeoutExpired:
        return {"error": f"Recalculation timed out after {timeout} seconds"}
    except FileNotFoundError:
        return {"error": "soffice (LibreOffice) not found. Please install LibreOffice."}

    if result.returncode != 0:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        else:
            return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}
