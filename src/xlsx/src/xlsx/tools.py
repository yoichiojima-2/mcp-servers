import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from . import mcp
from .recalc import recalc


def _read_excel_to_dataframe(file_path: str, sheet_name: str = "") -> pd.DataFrame:
    """Helper function to read Excel file into DataFrame with error handling.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name to read (default: first sheet)

    Returns:
        DataFrame containing the data

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is corrupted or invalid
        PermissionError: If file can't be accessed
    """
    path = Path(file_path).expanduser().resolve()

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if not path.is_file():
        raise ValueError(f"Path is not a file: {path}")

    try:
        if sheet_name:
            return pd.read_excel(str(path), sheet_name=sheet_name)
        else:
            return pd.read_excel(str(path))
    except PermissionError as e:
        raise PermissionError(f"Permission denied: {path}") from e
    except Exception as e:
        raise ValueError(f"Failed to read Excel file {path}: {str(e)}") from e


@mcp.tool()
def read_excel(file_path: str, sheet_name: str = "") -> str:
    """Read an Excel file and return its contents as markdown table.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name to read (default: first sheet)

    Returns:
        Markdown formatted table of the data
    """
    try:
        df = _read_excel_to_dataframe(file_path, sheet_name)
        return df.to_markdown(index=False)
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"


@mcp.tool()
def create_excel(file_path: str, data: str, sheet_name: str = "Sheet1") -> str:
    """Create a new Excel file with data.

    Args:
        file_path: Path to output Excel file
        data: CSV formatted data to write
        sheet_name: Name of the sheet (default: 'Sheet1')

    Returns:
        Success message
    """
    from io import StringIO

    try:
        path = Path(file_path).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)

        df = pd.read_csv(StringIO(data))
        df.to_excel(str(path), sheet_name=sheet_name, index=False)
        return f"Created {path} with {len(df)} rows"
    except Exception as e:
        return f"Error creating Excel file: {str(e)}"


@mcp.tool()
def write_cell(file_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Write a value or formula to a specific cell.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet
        cell: Cell address (e.g., 'A1', 'B5')
        value: Value or formula to write (formulas start with '=')

    Returns:
        Success message
    """
    try:
        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return f"Error: File not found: {path}"

        wb = load_workbook(str(path))
        try:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' not found in {path}"

            ws = wb[sheet_name]
            ws[cell] = value
            wb.save(str(path))
            return f"Wrote '{value}' to {sheet_name}!{cell}"
        finally:
            wb.close()
    except Exception as e:
        return f"Error writing to cell: {str(e)}"


@mcp.tool()
def recalculate(file_path: str, timeout: int = 30) -> str:
    """Recalculate all formulas in Excel file and check for errors.

    Args:
        file_path: Path to Excel file
        timeout: Maximum time to wait for recalculation (default: 30 seconds)

    Returns:
        JSON formatted result with error details
    """
    try:
        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return json.dumps({"error": f"File not found: {path}"}, indent=2)

        result = recalc(str(path), timeout)
        return json.dumps(result, indent=2)
    except Exception as e:
        return json.dumps({"error": f"Recalculation failed: {str(e)}"}, indent=2)


@mcp.tool()
def get_sheet_names(file_path: str) -> str:
    """Get list of sheet names in an Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        Comma-separated list of sheet names
    """
    try:
        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return f"Error: File not found: {path}"

        wb = load_workbook(str(path), read_only=True)
        try:
            sheets = ", ".join(wb.sheetnames)
            return sheets
        finally:
            wb.close()
    except Exception as e:
        return f"Error reading sheet names: {str(e)}"


@mcp.tool()
def add_sheet(file_path: str, sheet_name: str) -> str:
    """Add a new sheet to an Excel file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name for the new sheet

    Returns:
        Success message
    """
    try:
        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return f"Error: File not found: {path}"

        wb = load_workbook(str(path))
        try:
            if sheet_name in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' already exists in {path}"

            wb.create_sheet(sheet_name)
            wb.save(str(path))
            return f"Added sheet '{sheet_name}' to {path}"
        finally:
            wb.close()
    except Exception as e:
        return f"Error adding sheet: {str(e)}"


@mcp.tool()
def convert_to_csv(file_path: str, output_file: str, sheet_name: str = "") -> str:
    """Convert an Excel sheet to CSV.

    Args:
        file_path: Path to Excel file
        output_file: Path to output CSV file
        sheet_name: Sheet name to convert (default: first sheet)

    Returns:
        Success message
    """
    try:
        df = _read_excel_to_dataframe(file_path, sheet_name)
        output_path = Path(output_file).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(str(output_path), index=False)
        return f"Converted {file_path} to {output_path}"
    except Exception as e:
        return f"Error converting to CSV: {str(e)}"
